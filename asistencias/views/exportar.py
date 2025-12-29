from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404
from django.utils.timezone import localtime
from django.utils import timezone
from django.contrib.auth import get_user_model
from io import BytesIO
from openpyxl import Workbook

from ..models import (
    Diplomatura, Materia, Clase, Asistencia,
    ProfesorMateria, InscripcionDiplomatura, InscripcionMateria
)

def _dt(v):
    """Formatea datetimes/fechas a texto legible (local)."""
    if v is None:
        return ""
    if hasattr(v, "tzinfo"):
        return localtime(v).strftime("%Y-%m-%d %H:%M:%S")
    return v.strftime("%Y-%m-%d")

from openpyxl.utils import get_column_letter

def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        # Use get_column_letter instead of accessing column_letter directly on the cell,
        # which might be a MergedCell
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for r in rows:
        ws.append(r)
    _autosize(ws)

def exportar_xlsx(request):
    # Solo Coordinadores (3) o Administradores (5)
    if not request.user.is_authenticated or request.user.nivel not in (3, 5):
        return HttpResponseForbidden("No autorizado.")

    User = get_user_model()

    wb = Workbook()
    wb.remove(wb.active)

    # === Usuarios ===
    ws = wb.create_sheet("Usuarios")
    headers = [
        "id", "email", "first_name", "second_name", "last_name", "second_last_name",
        "dni", "nivel", "is_active", "date_joined", "last_login",
    ]
    rows = []
    for u in User.objects.all().order_by("last_name", "first_name"):
        rows.append([
            u.id, u.email, u.first_name, getattr(u, "second_name", ""),
            u.last_name, getattr(u, "second_last_name", ""),
            u.dni, u.nivel, u.is_active, _dt(u.date_joined), _dt(u.last_login)
        ])
    _write_sheet(ws, headers, rows)

    # === Diplomaturas ===
    ws = wb.create_sheet("Diplomaturas")
    headers = ["id", "nombre", "descripcion", "codigo", "creada_por_email", "coordinadores_emails"]
    rows = []
    for d in Diplomatura.objects.all().order_by("nombre"):
        coords = ", ".join(d.coordinadores.values_list("email", flat=True))
        rows.append([
            d.id, d.nombre, d.descripcion, d.codigo,
            d.creada_por.email if d.creada_por else "",
            coords
        ])
    _write_sheet(ws, headers, rows)

    # === Materias ===
    ws = wb.create_sheet("Materias")
    headers = ["id", "diplomatura_id", "diplomatura", "nombre", "descripcion", "codigo",
               "profesor_titular_id", "profesor_titular_email"]
    rows = []
    for m in Materia.objects.select_related("diplomatura", "profesor_titular").all().order_by("diplomatura__nombre", "nombre"):
        rows.append([
            m.id, m.diplomatura_id, m.diplomatura.nombre,
            m.nombre, m.descripcion, m.codigo,
            m.profesor_titular_id or "",
            m.profesor_titular.email if m.profesor_titular else ""
        ])
    _write_sheet(ws, headers, rows)

    # === Clases ===
    ws = wb.create_sheet("Clases")
    headers = ["id", "materia_id", "materia", "fecha", "hora_inicio", "hora_fin", "tema", "ventana_activa"]
    rows = []
    now = timezone.now()
    for c in Clase.objects.select_related("materia", "materia__diplomatura").all().order_by("-fecha"):
        ventana_activa = (c.hora_inicio <= now <= c.hora_fin)
        rows.append([
            c.id, c.materia_id, f"{c.materia.nombre} ({c.materia.diplomatura.nombre})",
            _dt(c.fecha), _dt(c.hora_inicio), _dt(c.hora_fin), c.tema, ventana_activa
        ])
    _write_sheet(ws, headers, rows)

    # === Asistencias ===
    ws = wb.create_sheet("Asistencias")
    headers = ["id", "clase_id", "materia", "fecha_clase",
               "user_id", "email_user", "dni_user", "presente", "timestamp"]
    rows = []
    asist_qs = Asistencia.objects.select_related(
        "clase", "clase__materia", "clase__materia__diplomatura", "user"
    ).all().order_by("-timestamp")
    for a in asist_qs:
        rows.append([
            a.id, a.clase_id,
            f"{a.clase.materia.nombre} ({a.clase.materia.diplomatura.nombre})",
            _dt(a.clase.fecha),
            a.user_id, a.user.email, a.user.dni,
            a.presente, _dt(a.timestamp)
        ])
    _write_sheet(ws, headers, rows)

    # === ProfesorMateria ===
    ws = wb.create_sheet("ProfesorMateria")
    headers = ["id", "user_id", "email", "materia_id", "materia", "diplomatura", "rol"]
    rows = []
    for pm in ProfesorMateria.objects.select_related("user", "materia", "materia__diplomatura").all():
        rows.append([
            pm.id, pm.user_id, pm.user.email,
            pm.materia_id, pm.materia.nombre, pm.materia.diplomatura.nombre, pm.rol
        ])
    _write_sheet(ws, headers, rows)

    # === InscripcionDiplomatura ===
    ws = wb.create_sheet("InscDiplomatura")
    headers = ["id", "user_id", "email", "dni", "diplomatura_id", "diplomatura", "fecha"]
    rows = []
    for ins in InscripcionDiplomatura.objects.select_related("user", "diplomatura").all():
        rows.append([
            ins.id, ins.user_id, ins.user.email, ins.user.dni,
            ins.diplomatura_id, ins.diplomatura.nombre, _dt(ins.fecha)
        ])
    _write_sheet(ws, headers, rows)

    # === InscripcionMateria ===
    ws = wb.create_sheet("InscMateria")
    headers = ["id", "user_id", "email", "dni", "materia_id", "materia", "diplomatura", "fecha"]
    rows = []
    for ins in InscripcionMateria.objects.select_related("user", "materia", "materia__diplomatura").all():
        rows.append([
            ins.id, ins.user_id, ins.user.email, ins.user.dni,
            ins.materia_id, ins.materia.nombre, ins.materia.diplomatura.nombre, _dt(ins.fecha)
        ])
    _write_sheet(ws, headers, rows)

    # ⚠️ No se exportan tokens para niveles < 5
    # (Si quisieras incluirlos solo para admin, podrías hacer un if request.user.nivel == 5:)

    # Respuesta HTTP
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"asistencias_export_{localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def exportar_asistencia_materia(request, materia_id):
    """
    Exporta una planilla de asistencia para una materia específica.
    Permisos: Nivel >= 3 (Coordinador) o Nivel 2 (Docente) si es titular o adjunto de la materia.
    """
    if not request.user.is_authenticated:
        return HttpResponseForbidden("No autenticado.")

    # Obtener materia o 404
    try:
        materia = Materia.objects.get(pk=materia_id)
    except Materia.DoesNotExist:
        return HttpResponseForbidden("Materia no encontrada.")

    # Verificación de permisos
    tiene_permiso = False
    if request.user.nivel >= 3:
        tiene_permiso = True
    elif request.user.nivel == 6: # Referente Municipal
        # Check if registered in the diplomatura of this materia
        if InscripcionDiplomatura.objects.filter(user=request.user, diplomatura=materia.diplomatura).exists():
            tiene_permiso = True
    elif request.user.nivel == 2:
        # Verificar si es titular
        if materia.profesor_titular_id == request.user.id:
            tiene_permiso = True
        else:
            # Verificar si es adjunto
            es_adjunto = ProfesorMateria.objects.filter(user=request.user, materia=materia).exists()
            if es_adjunto:
                tiene_permiso = True
    
    if not tiene_permiso:
        return HttpResponseForbidden("No tiene permisos para exportar asistencia de esta materia.")

    # Obtener Clases
    clases = Clase.objects.filter(materia=materia).order_by('fecha')
    
    # Obtener Alumnos inscriptos
    inscripciones = InscripcionMateria.objects.filter(materia=materia).select_related('user').order_by('user__last_name', 'user__first_name')
    alumnos = [i.user for i in inscripciones]

    # Obtener Asistencias
    # Diccionario: {(user_id, clase_id): presente (bool)}
    asistencias_qs = Asistencia.objects.filter(clase__materia=materia)
    asistencias_map = {(a.user_id, a.clase_id): a.presente for a in asistencias_qs}

    # Generar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(clases) + 2)
    ws.cell(row=1, column=1, value=f"Materia: {materia.nombre} - Diplomatura: {materia.diplomatura.nombre}")

    # Encabezados
    headers = ["Alumno"] + [_dt(c.fecha) for c in clases]
    ws.append(headers)

    # Filas de alumnos
    for alumno in alumnos:
        row = [f"{alumno.last_name}, {alumno.first_name}"]
        for clase in clases:
            presente = asistencias_map.get((alumno.id, clase.id))
            if presente is True:
                val = "P"
            elif presente is False:
                val = "A"
            else:
                val = "-" # No hay registro (ausente o no tomada)
            row.append(val)
        ws.append(row)

    _autosize(ws)

    # Respuesta HTTP
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"asistencia_{materia.codigo}_{localtime(timezone.now()).strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def exportar_asistencia_diplomatura(request, diplomatura_id):
    """
    Exporta la asistencia de TODAS las materias de una diplomatura.
    Cada materia en una hoja distinta.
    """
    if not request.user.is_authenticated:
        return HttpResponseForbidden("No autenticado.")

    diplomatura = get_object_or_404(Diplomatura, id=diplomatura_id)

    # Permisos: Coordinador (3), Admin (5), Referente (6) (si está inscripto)
    tiene_permiso = False
    if request.user.nivel in [3, 5, 7]: # Incluye Supervisor
        tiene_permiso = True
    elif request.user.nivel == 6:
        if InscripcionDiplomatura.objects.filter(user=request.user, diplomatura=diplomatura).exists():
            tiene_permiso = True
            
    if not tiene_permiso:
        return HttpResponseForbidden("No tenés permiso para exportar esta diplomatura.")

    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet

    materias = Materia.objects.filter(diplomatura=diplomatura).order_by('nombre')
    
    if not materias.exists():
         ws = wb.create_sheet("Info")
         ws.append(["No hay materias en esta diplomatura."])

    for materia in materias:
        # Create sheet for materia (limit name length to 31 chars)
        sheet_name = materia.nombre[:30]
        ws = wb.create_sheet(sheet_name)
        
        clases = Clase.objects.filter(materia=materia).order_by('fecha')
        inscripciones = InscripcionMateria.objects.filter(materia=materia).select_related('user').order_by('user__last_name', 'user__first_name')
        alumnos = [i.user for i in inscripciones]
        
        asistencias_qs = Asistencia.objects.filter(clase__materia=materia)
        asistencias_map = {(a.user_id, a.clase_id): a.presente for a in asistencias_qs}

        # Header
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(clases) + 2)
        ws.cell(row=1, column=1, value=f"{materia.nombre}")
        
        headers = ["Alumno"] + [_dt(c.fecha) for c in clases]
        ws.append(headers)
        
        for alumno in alumnos:
            row = [f"{alumno.last_name}, {alumno.first_name}"]
            for clase in clases:
                presente = asistencias_map.get((alumno.id, clase.id))
                val = "P" if presente is True else ("A" if presente is False else "-")
                row.append(val)
            ws.append(row)
            
        _autosize(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"asistencia_diplomatura_{diplomatura.codigo}_{localtime(timezone.now()).strftime('%Y%m%d')}.xlsx"
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
