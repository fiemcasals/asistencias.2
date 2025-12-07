from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.utils import timezone
from asistencias.models import Diplomatura, Materia, Clase, Asistencia, InscripcionMateria, ProfesorMateria
import openpyxl
from io import BytesIO
import datetime

User = get_user_model()

class ExportarAsistenciaTest(TestCase):
    def setUp(self):
        self.client = Client()
        
        # Crear usuarios
        self.admin = User.objects.create_superuser(email='admin@test.com', password='password', first_name='Admin', last_name='User', dni='1')
        self.coordinador = User.objects.create_user(email='coord@test.com', password='password', first_name='Coord', last_name='User', dni='2', nivel=3)
        self.profesor_titular = User.objects.create_user(email='profe@test.com', password='password', first_name='Profe', last_name='Titular', dni='3', nivel=2)
        self.profesor_adjunto = User.objects.create_user(email='adjunto@test.com', password='password', first_name='Profe', last_name='Adjunto', dni='4', nivel=2)
        self.profesor_otro = User.objects.create_user(email='otro@test.com', password='password', first_name='Profe', last_name='Otro', dni='5', nivel=2)
        self.alumno = User.objects.create_user(email='alumno@test.com', password='password', first_name='Alumno', last_name='Uno', dni='6', nivel=1)
        
        # Crear estructura académica
        self.diplomatura = Diplomatura.objects.create(nombre='Diplo Test', codigo='D1', creada_por=self.admin)
        self.materia = Materia.objects.create(diplomatura=self.diplomatura, nombre='Materia Test', codigo='M1', profesor_titular=self.profesor_titular)
        
        # Asignar adjunto
        ProfesorMateria.objects.create(user=self.profesor_adjunto, materia=self.materia, rol='adjunto')
        
        # Inscribir alumno
        InscripcionMateria.objects.create(user=self.alumno, materia=self.materia)
        
        # Crear clases
        self.clase1 = Clase.objects.create(materia=self.materia, fecha=datetime.date(2023, 10, 1), hora_inicio=timezone.now(), hora_fin=timezone.now())
        self.clase2 = Clase.objects.create(materia=self.materia, fecha=datetime.date(2023, 10, 8), hora_inicio=timezone.now(), hora_fin=timezone.now())
        
        # Crear asistencia
        # Clase 1: Presente
        Asistencia.objects.create(clase=self.clase1, user=self.alumno, presente=True)
        # Clase 2: Ausente
        Asistencia.objects.create(clase=self.clase2, user=self.alumno, presente=False)
        
        self.url = reverse('asistencias:exportar_asistencia_materia', args=[self.materia.id])

    def test_acceso_anonimo(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 403)

    def test_acceso_alumno(self):
        self.client.force_login(self.alumno)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 403)

    def test_acceso_profesor_otro(self):
        self.client.force_login(self.profesor_otro)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 403)

    def test_acceso_profesor_titular(self):
        self.client.force_login(self.profesor_titular)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self._verificar_excel(response)

    def test_acceso_profesor_adjunto(self):
        self.client.force_login(self.profesor_adjunto)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self._verificar_excel(response)

    def test_acceso_coordinador(self):
        self.client.force_login(self.coordinador)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self._verificar_excel(response)

    def _verificar_excel(self, response):
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Verificar título
        self.assertIn('Materia: Materia Test', ws.cell(row=1, column=1).value)
        
        # Verificar encabezados (Alumno + 2 fechas)
        self.assertEqual(ws.cell(row=2, column=1).value, 'Alumno')
        # Las fechas pueden venir como string o datetime dependiendo de _dt, asumimos string YYYY-MM-DD
        self.assertIn('2023-10-01', str(ws.cell(row=2, column=2).value))
        self.assertIn('2023-10-08', str(ws.cell(row=2, column=3).value))
        
        # Verificar datos alumno
        self.assertEqual(ws.cell(row=3, column=1).value, 'Uno, Alumno')
        self.assertEqual(ws.cell(row=3, column=2).value, 'P')
        self.assertEqual(ws.cell(row=3, column=3).value, 'A')
